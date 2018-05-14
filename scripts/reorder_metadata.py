import argparse
from Bio import SeqIO
import openpyxl

parser = argparse.ArgumentParser(
    description='Reorder excel file based on first column and fasta seq id.')
parser.add_argument('input_fasta', type=str,
    help='Input fasta file')
parser.add_argument('input_xlsx', type=str,
    help='Input excel file')
parser.add_argument('output', type=str, nargs='?',
    default='information_ordered.xlsx',
    help='Output excel file (default: information_ordered.xlsx)')
args = parser.parse_args()

sequences = {}
wb = openpyxl.load_workbook(args.input_xlsx)
sheet = wb['Sheet']
for row in sheet.iter_rows():
    if row[0].value.split('/')[0].split('.')[0].split('-')[0] in sequences:
        print('ERROR: duplicates possibly')
    sequences[row[0].value.split('/')[0].split('.')[0].split('-')[0]] = row

wb_output = openpyxl.Workbook()
sheet_out = wb_output.active
row_num = 0
for sequence in SeqIO.parse(args.input_fasta, 'fasta'):
    if not (sequence.description.split('/')[0].split('.')[0].split('-')[0] in sequences):
        print("NOT FOUND in information input: " + str(sequence.description))
        continue
    for col_num in range(0, sheet.max_column):
        sheet_out.cell(row=(row_num + 1), column=(col_num + 1)).value = sequences[sequence.description.split('/')[0].split('.')[0].split('-')[0]][col_num].value
    row_num += 1
    if row_num % 100 == 0:
        print(str(row_num))

print()
wb_output.save(args.output)
