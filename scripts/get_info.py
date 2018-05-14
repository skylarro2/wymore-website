#!/usr/local/bin/python3
import requests                     # Make requests to API
import xml.etree.ElementTree as ET  # Parse XML response
                                    # NOTE: Has XML vulnerabilities (2/19/2018)
from Bio import SeqIO               # Read fasta files
import openpyxl                     # Write spreadsheets
import re                           # Searching with regex
import argparse                     # Parse arguments

base_url = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils'
base_url_emblebi = 'https://www.ebi.ac.uk/proteins/api'

def get_xml(url):
    response = requests.get(url)
    if response.status_code != 200:
        print('Response code was not 200')
        print(response.raw)
        return get_xml(url)
    return ET.fromstring(response.content)

def get_json(url):
    return requests.get(url).json()

# Returns '' if user says to skip
def get_search_term(description):
    result = re.search('^.*\{.*\|EMBL:(.*)\}', description)
    if not result is None:
        # use EMBL for search term
        return result.group(1) # first parenthesized subgroup
    else:
        # Use first part (space seperated)
        return description.split('/')[0].split('.')[0]
        """
        print('What should we search for: (leave blank to skip)')
        print(description)
        return input()
        """

# extract id from description
def get_seq_id(description):
    result = re.search('^(.*)\.\d\/.*$', description)
    return result.group(1)

def get_features(sequence_id):
    url = '{}/features?offset=0&size=100&accession={}'.format(base_url_emblebi,
                                                              sequence_id)
    data = get_json(url)
    if len(data) < 1:
        return None
    return data[0]

def get_taxonomy(taxid):
    url = '{}/taxonomy/id/{}'.format(base_url_emblebi, taxid)
    data = get_json(url)
    return data

# get sequence uid by searching term
def get_seq_uid(term):
    url = '{}/esearch.fcgi?db=protein&term={}&retmode=xml'.format(base_url, term)
    tree = get_xml(url)
    seq_uid = tree.find('IdList').find('Id')
    if seq_uid is None:
        return None
    return seq_uid.text

def fetch_data(uid):
    # fetch data
    url = '{}/efetch.fcgi?db=protein&id={}&retmode=xml'.format(base_url, uid)
    tree = get_xml(url)
    return tree

def find_organism(tree):
    organism = tree.find('GBSeq') \
                    .find('GBSeq_organism').text
    # tree.find('Entrezgene') \
    #                 .find('Entrezgene_source') \
    #                 .find('BioSource') \
    #                 .find('BioSource_org') \
    #                 .find('Org-ref') \
    #                 .find('Org-ref_taxname').text
    return organism

def find_product(tree):
    GBFeatures = tree.find('GBSeq') \
                        .find('GBSeq_feature-table') \
                        .findall('GBFeature')
    for f in GBFeatures:
        quals = f.find('GBFeature_quals').findall('GBQualifier')
        for q in quals:
            if q.find('GBQualifier_name').text == 'product':
                return q.find('GBQualifier_value').text
    # tree.find('Entrezgene') \
    #                 .find('Entrezgene_prot') \
    #                 .find('Prot-ref') \
    #                 .find('Prot-ref_name') \
    #                 .find('Prot-ref_name_E').text
    print('Product not found')
    return '' # product

def find_product_note(tree):
    """ Kind of redundant with find_product. TODO: combine these functions """
    GBFeatures = tree.find('GBSeq') \
                        .find('GBSeq_feature-table') \
                        .findall('GBFeature')
    for f in GBFeatures:
        quals = f.find('GBFeature_quals').findall('GBQualifier')
        for q in quals:
            if q.find('GBQualifier_name').text == 'product':
                for m in quals:
                    if m.find('GBQualifier_name').text == 'note':
                        return m.find('GBQualifier_value').text
    print('Note not found')
    return ''

def search_ncbi(sequence):
    term = get_search_term(sequence.description)
    if term == '':
        return {}
    uid = get_seq_uid(term)
    if uid is None:
        print('Not found: ', sequence.description)
        return {'term': term}
    data = fetch_data(uid)
    return {'term': term,
            'organism': find_organism(data),
            'product': find_product(data),
            'notes': find_product_note(data)}

def search_emblebi(sequence):
    ret = {}
    ret['sequence_id'] = get_seq_id(sequence.description)
    if ret['sequence_id'] == '':
        return ret
    # Search for features
    features = get_features(ret['sequence_id'])
    if features is None:
        ret['no_features'] = 'No features found'
        return ret
    if features['sequence'] != str(sequence.seq).replace('-',''):
        print('ERROR, sequences do not match on: {}'.format(str(sequence)))
        print('EMBL-EBI response: {}'.format(features))
    ret['taxid'] = features.get('taxid')
    ret['accession'] = features.get('accession')
    ret['entryName'] = features.get('entryName')
    if len(features['features']) > 0:
        ret['description'] = features['features'][0].get('description')
        ret['evidenced'] = features['features'][0].get('evidences')
    if ret['taxid'] is None:
        return ret
    # Search using taxid and merge with result
    ret = {**ret, **get_taxonomy(ret['taxid'])}
    return ret

def main():
    # parse arguments
    parser = argparse.ArgumentParser(
        description='Gather information about each sequence.')
    parser.add_argument('input', type=str,
        help='Input fasta file')
    parser.add_argument('output', type=str, nargs='?',
        default='information.xlsx',
        help='Output excel file (default: information.xlsx)')
    parser.add_argument('-n', '--ncbi', help='Search NCBI database',
                    action='store_true')
    parser.add_argument('-e', '--emblebi', help='Search EMBL-EBI database',
                    action='store_true')
    parser.add_argument('-v', '--verbose', help='Verbose output while running',
                    action='store_true')
    parser.add_argument('-s', '--start', type=int, nargs='?',
        default=0,
        help='Starts at this index in the input file. (default is 0)')
    parser.add_argument('-a', '--save', type=int, nargs='?',
        default=100,
        help='Number of searches between each excel file save. (default is 100)')
    args = parser.parse_args()
    wb = openpyxl.Workbook()
    sheet = wb.active
    # Write headers
    sheet.cell(row=1, column=1).value = 'Description'
    sheet_row = 1
    sequence_number = 0
    column_names = {}
    sheet_column = 2
    for sequence in SeqIO.parse(args.input, 'fasta'):
        sequence_number += 1
        if args.start > sequence_number:
            continue
        if args.verbose:
            print('[{}] {}'.format(sheet_row, sequence.name))
        sheet_row += 1
        sheet.cell(row=sheet_row, column=1).value = sequence.description
        result = {}
        if args.ncbi:
            result = search_ncbi(sequence)
        if args.emblebi:
            result = search_emblebi(sequence)
            if args.verbose:
                print('\t\tTax id: {}'.format(result.get('taxid', 'not found')), end='')
        for key, value in result.items():
            if not (key in column_names):
                column_names[key] = sheet_column
                sheet_column += 1
            sheet.cell(row=sheet_row, column=column_names[key]).value = str(value)
        # Only save every 100 rows
        if sheet_row % args.save == 0:
            # Add column_names as headers
            print(column_names)
            for k, v in column_names.items():
                sheet.cell(row=1, column=v).value = str(k)
            if args.verbose:
                print('Saving...', end='\r')
            wb.save(args.output)
            print('Saved. Continuing...')

    # Save end result
    wb.save(args.output)
    print('Saved.')
if __name__ == '__main__':
    main()
